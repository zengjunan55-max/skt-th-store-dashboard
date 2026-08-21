import json
import os
import shutil
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from urllib.request import urlopen

import openpyxl


SHEET_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "142eo8ASRHj0utna0cHxiWRnMY8s7r0ZFKAc-7s_jdgs/export?format=xlsx"
)
WORKSPACE_ROOT = Path(r"C:\Users\Administrator\Desktop\AI-shopee")
SOURCE_XLSX = WORKSPACE_ROOT / "sheet-latest.xlsx"
SITE_ROOT = WORKSPACE_ROOT / "site"
STORE_OUTPUT = SITE_ROOT / "store-trend-data.js"
PRODUCT_OUTPUT = SITE_ROOT / "product-trend-data.js"


def norm(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def num(value):
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def recognized(value):
    text = norm(value)
    return bool(text) and text != "#N/A"


def download_latest_workbook():
    SOURCE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with urlopen(SHEET_EXPORT_URL, timeout=60) as response:
        payload = response.read()
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
        handle.write(payload)
        temp_path = Path(handle.name)
    try:
        temp_path.replace(SOURCE_XLSX)
        return SOURCE_XLSX, None
    except PermissionError:
        # If the local workbook is open in Excel, keep using the fresh temp file.
        return temp_path, str(SOURCE_XLSX)


def build_sheet(ws, builder):
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers) if header is not None}
    data = []
    for row in rows:
        item = builder(row, idx)
        if item:
            data.append(item)
    return data


def write_js(output_path, variable_name, payload):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        handle.write(f"window.{variable_name} = ")
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
        handle.write(";\n")


def build_payloads(source_path):
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

    store_ws = wb["店铺基础数据"]
    inner_ws = wb["站内广告"]
    all_ws = wb["ALL_data"]

    category_by_product = {}
    category_store_map = {}
    product_store_map = {}

    def register_category(product, category, store=""):
        product_name = norm(product)
        category_name = norm(category)
        store_name = norm(store)
        if not recognized(product_name) or not recognized(category_name):
            return
        category_by_product.setdefault(product_name, category_name)
        if store_name:
            category_store_map[(store_name, product_name)] = category_name

    def build_store_main(row, idx):
        category = norm(row[idx.get("品类")])
        product = norm(row[idx.get("商品名称")])
        store = norm(row[idx.get("店铺")])
        register_category(product, category, store)
        if recognized(product) and recognized(store):
            product_store_map[(store, product)] = product
        return {
            "date": norm(row[idx.get("date")]),
            "store": store,
            "category": category,
            "sales_thb": num(row[idx.get("Sales (Confirmed Order) (THB)")]),
            "visitors": num(row[idx.get("Product Visitors (Visit)")]),
            "buyers": num(row[idx.get("Buyers (Confirmed Order)")]),
            "units": num(row[idx.get("Units (Confirmed Order)")]),
            "total_purchase_people": num(row[idx.get("总购买人数")]),
            "new_customer": num(row[idx.get("旗舰店新客")]),
            "old_customer": num(row[idx.get("旗舰店老客")]),
        }

    def build_product_main(row, idx):
        product = norm(row[idx.get("商品名称")])
        if not recognized(product):
            return None
        category = norm(row[idx.get("品类")])
        store = norm(row[idx.get("店铺")])
        register_category(product, category, store)
        if recognized(store):
            product_store_map[(store, product)] = product
        return {
            "date": norm(row[idx.get("date")]),
            "store": store,
            "category": category,
            "product": product,
            "sales_thb": num(row[idx.get("Sales (Confirmed Order) (THB)")]),
            "visitors": num(row[idx.get("Product Visitors (Visit)")]),
            "buyers": num(row[idx.get("Buyers (Confirmed Order)")]),
            "units": num(row[idx.get("Units (Confirmed Order)")]),
            "total_purchase_people": num(row[idx.get("总购买人数")]),
            "new_customer": num(row[idx.get("旗舰店新客")]),
            "old_customer": num(row[idx.get("旗舰店老客")]),
        }

    def build_store_inner(row, idx):
        return {
            "date": norm(row[idx.get("date")]),
            "store": norm(row[idx.get("店铺")]),
            "category": norm(row[idx.get("类目")]),
            "ads_spend_thb": num(row[idx.get("Ads Spend(THB)")]),
            "spend_rmb": num(row[idx.get("Spend-人民币")]),
            "clicks": num(row[idx.get("Clicks")]),
            "cpc": num(row[idx.get("CPC")]),
            "gmv_thb": num(row[idx.get("GMV")]),
        }

    def build_product_inner(row, idx):
        product = norm(row[idx.get("站内产品命名")])
        if not recognized(product):
            return None
        category = norm(row[idx.get("类目")])
        store = norm(row[idx.get("店铺")])
        register_category(product, category, store)
        if recognized(store):
            product_store_map[(store, product)] = product
        return {
            "date": norm(row[idx.get("date")]),
            "store": store,
            "category": category,
            "product": product,
            "ads_spend_thb": num(row[idx.get("Ads Spend(THB)")]),
            "spend_rmb": num(row[idx.get("Spend-人民币")]),
            "clicks": num(row[idx.get("Clicks")]),
            "cpc": num(row[idx.get("CPC")]),
            "gmv_thb": num(row[idx.get("GMV")]),
        }

    store_main_rows = build_sheet(store_ws, build_store_main)
    product_main_rows = build_sheet(store_ws, build_product_main)
    store_inner_rows = build_sheet(inner_ws, build_store_inner)
    product_inner_rows = build_sheet(inner_ws, build_product_inner)

    rows = all_ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers) if header is not None}

    store_outer_rows = []
    store_brand_rows = []
    product_outer_rows = []
    product_brand_rows = []

    for row in rows:
        date_value = norm(row[idx.get("日期")])
        category_value = norm(row[idx.get("品类")]) if idx.get("品类") is not None else ""
        product = norm(row[idx.get("产品")]) if idx.get("产品") is not None else ""
        if not recognized(product):
            product = norm(row[idx.get("Product")]) if idx.get("Product") is not None else ""
        if not recognized(product):
            continue

        store = norm(row[idx.get("店铺")]) if idx.get("店铺") is not None else ""
        category = category_value or category_store_map.get((store, product), "") or category_by_product.get(product, "")
        spend_usd = num(row[idx.get("花费金额（USD）")])
        impressions = num(row[idx.get("展示次数")])
        clicks = num(row[idx.get("点击量")])
        conversion_value = num(row[idx.get("购物转化价值")])
        orders = num(row[idx.get("订单数")])
        pitcher = norm(row[idx.get("投手")])
        ad_type = norm(row[idx.get("广告类型")])
        ad_form2 = norm(row[idx.get("广告形式2")])
        store_item = {
            "date": date_value,
            "store": store,
            "category": category,
            "spend_usd": spend_usd,
            "impressions": impressions,
            "clicks": clicks,
            "conversion_value": conversion_value,
            "orders": orders,
            "ad_type": ad_type,
            "ad_form2": ad_form2,
        }
        product_item = {
            "date": date_value,
            "store": store,
            "category": category,
            "product": product_store_map.get((store, product), product),
            "spend_usd": spend_usd,
            "impressions": impressions,
            "clicks": clicks,
            "conversion_value": conversion_value,
            "orders": orders,
            "ad_type": ad_type,
            "ad_form2": ad_form2,
        }

        if pitcher == "SKT":
            store_brand_rows.append(store_item)
            product_brand_rows.append(product_item)
        else:
            store_outer_rows.append(store_item)
            product_outer_rows.append(product_item)

    store_payload = {
        "stores": sorted({row["store"] for row in store_main_rows if row.get("store")}),
        "main": store_main_rows,
        "inner": store_inner_rows,
        "outer": store_outer_rows,
        "brand": store_brand_rows,
    }
    product_payload = {
        "main": product_main_rows,
        "inner": product_inner_rows,
        "outer": product_outer_rows,
        "brand": product_brand_rows,
    }
    return store_payload, product_payload


def summarize_dates(payload):
    summary = {}
    for key in ("main", "inner", "outer", "brand"):
        rows = payload.get(key, [])
        dates = sorted({row.get("date") for row in rows if row.get("date")})
        summary[key] = {
            "count": len(rows),
            "min": dates[0] if dates else None,
            "max": dates[-1] if dates else None,
        }
    return summary


def run_refresh():
    source_path, locked_target = download_latest_workbook()
    try:
        store_payload, product_payload = build_payloads(source_path)
        write_js(STORE_OUTPUT, "storeTrendData", store_payload)
        write_js(PRODUCT_OUTPUT, "productTrendData", product_payload)
        result = {
            "source": str(source_path),
            "store": summarize_dates(store_payload),
            "product": summarize_dates(product_payload),
        }
        if locked_target:
            result["warning"] = f"本地文件占用，已跳过覆盖：{locked_target}"
        return result
    finally:
        if source_path != SOURCE_XLSX and source_path.exists():
            try:
                os.remove(source_path)
            except OSError:
                pass


def main():
    result = run_refresh()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
