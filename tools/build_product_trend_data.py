import json
from datetime import date, datetime

import openpyxl


SOURCE = r"C:\Users\Administrator\Desktop\AI-shopee\sheet-latest.xlsx"
OUTPUT = r"C:\Users\Administrator\Desktop\AI-shopee\site\product-trend-data.js"


def norm(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def num(value):
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def recognized(value):
    text = norm(value)
    return bool(text) and text != "#N/A"


def build_sheet(ws, builder):
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers) if header is not None}
    data = []
    for row in rows:
        item = builder(row, idx)
        if item:
            data.append(item)
    return data


def register_category(mapping, product, category):
    product_name = norm(product)
    category_name = norm(category)
    if not recognized(product_name) or not recognized(category_name):
        return
    mapping.setdefault(product_name, category_name)


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)

    store_ws = wb["店铺基础数据"]
    inner_ws = wb["站内广告"]
    all_ws = wb["ALL_data"]

    category_by_product = {}

    def build_main(row, idx):
        product = norm(row[idx.get("商品名称")])
        if not recognized(product):
            return None
        category = norm(row[idx.get("品类")])
        register_category(category_by_product, product, category)
        return {
            "date": norm(row[idx.get("date")]),
            "store": norm(row[idx.get("店铺")]),
            "category": category,
            "product": product,
            "sales_thb": num(row[idx.get("Sales (Confirmed Order) (THB)")]),
            "visitors": num(row[idx.get("Product Visitors (Visit)")]),
            "buyers": num(row[idx.get("Buyers (Confirmed Order)")]),
            "units": num(row[idx.get("Units (Confirmed Order)")]),
        }

    def build_inner(row, idx):
        product = norm(row[idx.get("站内产品命名")])
        if not recognized(product):
            return None
        category = norm(row[idx.get("类目")])
        register_category(category_by_product, product, category)
        return {
            "date": norm(row[idx.get("date")]),
            "store": norm(row[idx.get("店铺")]),
            "category": category,
            "product": product,
            "ads_spend_thb": num(row[idx.get("Ads Spend(THB)")]),
            "spend_rmb": num(row[idx.get("Spend-人民币")]),
            "clicks": num(row[idx.get("Clicks")]),
            "cpc": num(row[idx.get("CPC")]),
            "gmv_thb": num(row[idx.get("GMV")]),
        }

    main_rows = build_sheet(store_ws, build_main)
    inner_rows = build_sheet(inner_ws, build_inner)

    outer_rows = []
    brand_rows = []

    rows = all_ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers) if header is not None}

    for row in rows:
        product = norm(row[idx.get("产品")]) if idx.get("产品") is not None else ""
        if not recognized(product):
            product = norm(row[idx.get("Product")]) if idx.get("Product") is not None else ""
        if not recognized(product):
            continue

        item = {
            "date": norm(row[idx.get("日期")]),
            "category": category_by_product.get(product, ""),
            "product": product,
            "spend_usd": num(row[idx.get("花费金额（USD）")]),
        }
        pitcher = norm(row[idx.get("投手")])
        if pitcher == "SKT":
            brand_rows.append(item)
        else:
            outer_rows.append(item)

    payload = {
        "main": main_rows,
        "inner": inner_rows,
        "outer": outer_rows,
        "brand": brand_rows,
    }

    with open(OUTPUT, "w", encoding="utf-8") as handle:
        handle.write("window.productTrendData = ")
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
        handle.write(";\n")

    print(f"wrote {OUTPUT}")
    print(
        f"main={len(main_rows)} inner={len(inner_rows)} "
        f"outer={len(outer_rows)} brand={len(brand_rows)} "
        f"mapped_categories={len(category_by_product)}"
    )


if __name__ == "__main__":
    main()
