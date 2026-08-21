import json
import time
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from urllib.request import urlopen
from http.client import IncompleteRead

import openpyxl


SHEET_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "142eo8ASRHj0utna0cHxiWRnMY8s7r0ZFKAc-7s_jdgs/export?format=xlsx"
)
DEFAULT_REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_XLSX = DEFAULT_REPO_ROOT / ".cache" / "sheet-latest.xlsx"
LOCAL_WINDOWS_REPO_ROOT = Path(r"C:\Users\Administrator\Desktop\AI-shopee\github-pages-store-dashboard")
LOCAL_WINDOWS_SOURCE_XLSX = Path(r"C:\Users\Administrator\Desktop\AI-shopee\sheet-latest.xlsx")
REPO_ROOT = LOCAL_WINDOWS_REPO_ROOT if LOCAL_WINDOWS_REPO_ROOT.exists() else DEFAULT_REPO_ROOT
SOURCE_XLSX = LOCAL_WINDOWS_SOURCE_XLSX if LOCAL_WINDOWS_REPO_ROOT.exists() else DEFAULT_SOURCE_XLSX


def norm(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def num(value):
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def recognized(value):
    text = norm(value)
    return bool(text) and text != "#N/A"


def download_latest_workbook():
    SOURCE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    last_error = None
    for attempt in range(3):
        try:
            with urlopen(SHEET_EXPORT_URL, timeout=120) as response:
                SOURCE_XLSX.write_bytes(response.read())
            return SOURCE_XLSX
        except IncompleteRead as exc:
            last_error = exc
            if attempt == 2:
                raise
            time.sleep(2)
    if last_error:
        raise last_error
    return SOURCE_XLSX


def build_sheet(ws, builder):
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers) if header is not None}
    data = []
    for row in rows:
        item = builder(row, idx)
        if item:
            data.append(item)
    return data


def get_cell(row, idx, header, default=""):
    position = idx.get(header)
    if position is None or position >= len(row):
        return default
    return row[position]


def write_assignment(path, prelude, target, payload):
    with path.open("w", encoding="utf-8") as handle:
        if prelude:
            handle.write(prelude)
            if not prelude.endswith("\n"):
                handle.write("\n")
        handle.write(f"{target} = ")
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
        handle.write(";\n")


def split_rows_by_size(rows, prelude, target, max_bytes):
    chunks = []
    current_rows = []
    base_prefix = ""
    if prelude:
        base_prefix = prelude if prelude.endswith("\n") else f"{prelude}\n"
    base_prefix += f"{target} = {target} || [];\n{target}.push(..."
    base_suffix = ");\n"
    base_size = len((base_prefix + "[]" + base_suffix).encode("utf-8"))
    current_size = base_size

    for row in rows:
        row_json = json.dumps(row, ensure_ascii=False, separators=(",", ":"))
        row_size = len(row_json.encode("utf-8")) + (1 if current_rows else 0)
        if current_rows and current_size + row_size > max_bytes:
            chunks.append(current_rows)
            current_rows = [row]
            current_size = base_size + len(row_json.encode("utf-8"))
        else:
            current_rows.append(row)
            current_size += row_size

    if current_rows:
        chunks.append(current_rows)
    return chunks


def write_chunked_loader(path, prelude, target, rows, max_bytes=24 * 1024 * 1024):
    part_pattern = f"{path.stem}.part*.js"
    for old_part in path.parent.glob(part_pattern):
        old_part.unlink()

    chunks = split_rows_by_size(rows, prelude, target, max_bytes)
    part_files = []
    for index, chunk_rows in enumerate(chunks, start=1):
        part_path = path.with_name(f"{path.stem}.part{index}.js")
        part_files.append(f"./{part_path.name}")
        with part_path.open("w", encoding="utf-8") as handle:
            if prelude:
                handle.write(prelude)
                if not prelude.endswith("\n"):
                    handle.write("\n")
            handle.write(f"{target} = {target} || [];\n")
            handle.write(f"{target}.push(...")
            json.dump(chunk_rows, handle, ensure_ascii=False, separators=(",", ":"))
            handle.write(");\n")

    with path.open("w", encoding="utf-8") as handle:
        if prelude:
            handle.write(prelude)
            if not prelude.endswith("\n"):
                handle.write("\n")
        handle.write(f"{target} = {target} || [];\n")
        handle.write("window.__dashboardDataLoads = window.__dashboardDataLoads || [];\n")
        handle.write("window.__dashboardDataLoads.push((async () => {\n")
        handle.write("  const loadScript = src => new Promise((resolve, reject) => {\n")
        handle.write("    const script = document.createElement('script');\n")
        handle.write("    script.src = src;\n")
        handle.write("    script.onload = resolve;\n")
        handle.write("    script.onerror = () => reject(new Error(`Failed to load ${src}`));\n")
        handle.write("    document.head.appendChild(script);\n")
        handle.write("  });\n")
        for part_file in part_files:
            handle.write(f"  await loadScript('{part_file}');\n")
        handle.write("})());\n")


def load_workbook(path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def build_store_and_product_payloads(wb):
    store_ws = wb["店铺基础数据"]
    inner_ws = wb["站内广告"]
    all_ws = wb["ALL_data"]
    subsidy_ws = wb["补贴"]

    category_by_product = {}
    category_store_map = {}
    product_store_map = {}

    def register_category(product, category, store=""):
        product_name = norm(product)
        category_name = norm(category)
        store_name = norm(store)
        if not recognized(product_name) or not recognized(category_name):
            return
        category_by_product.setdefault(product_name, category_name)
        if store_name:
            category_store_map[(store_name, product_name)] = category_name

    def build_store_main(row, idx):
        category = norm(get_cell(row, idx, "品类"))
        product = norm(get_cell(row, idx, "商品名称"))
        store = norm(get_cell(row, idx, "店铺"))
        register_category(product, category, store)
        if recognized(product) and recognized(store):
            product_store_map[(store, product)] = product
        return {
            "date": norm(get_cell(row, idx, "date")),
            "store": store,
            "category": category,
            "sales_thb": num(get_cell(row, idx, "Sales (Confirmed Order) (THB)")),
            "visitors": num(get_cell(row, idx, "Product Visitors (Visit)")),
            "buyers": num(get_cell(row, idx, "Buyers (Confirmed Order)")),
            "units": num(get_cell(row, idx, "Units (Confirmed Order)")),
            "total_purchase_people": num(get_cell(row, idx, "总购买人数")),
            "new_customer": num(get_cell(row, idx, "旗舰店新客")),
            "old_customer": num(get_cell(row, idx, "旗舰店老客")),
        }

    def build_product_main(row, idx):
        product = norm(get_cell(row, idx, "商品名称"))
        if not recognized(product):
            return None
        category = norm(get_cell(row, idx, "品类"))
        store = norm(get_cell(row, idx, "店铺"))
        register_category(product, category, store)
        if recognized(store):
            product_store_map[(store, product)] = product
        return {
            "date": norm(get_cell(row, idx, "date")),
            "store": store,
            "category": category,
            "product": product,
            "sales_thb": num(get_cell(row, idx, "Sales (Confirmed Order) (THB)")),
            "visitors": num(get_cell(row, idx, "Product Visitors (Visit)")),
            "buyers": num(get_cell(row, idx, "Buyers (Confirmed Order)")),
            "units": num(get_cell(row, idx, "Units (Confirmed Order)")),
            "total_purchase_people": num(get_cell(row, idx, "总购买人数")),
            "new_customer": num(get_cell(row, idx, "旗舰店新客")),
            "old_customer": num(get_cell(row, idx, "旗舰店老客")),
        }

    def build_store_inner(row, idx):
        return {
            "date": norm(get_cell(row, idx, "date")),
            "store": norm(get_cell(row, idx, "店铺")),
            "category": norm(get_cell(row, idx, "类目")),
            "ads_spend_thb": num(get_cell(row, idx, "Ads Spend(THB)")),
            "spend_rmb": num(get_cell(row, idx, "Spend-人民币")),
            "clicks": num(get_cell(row, idx, "Clicks")),
            "cpc": num(get_cell(row, idx, "CPC")),
            "gmv_thb": num(get_cell(row, idx, "GMV")),
        }

    def build_product_inner(row, idx):
        product = norm(get_cell(row, idx, "站内产品命名"))
        if not recognized(product):
            return None
        category = norm(get_cell(row, idx, "类目"))
        store = norm(get_cell(row, idx, "店铺"))
        register_category(product, category, store)
        if recognized(store):
            product_store_map[(store, product)] = product
        return {
            "date": norm(get_cell(row, idx, "date")),
            "store": store,
            "category": category,
            "product": product,
            "ads_spend_thb": num(get_cell(row, idx, "Ads Spend(THB)")),
            "spend_rmb": num(get_cell(row, idx, "Spend-人民币")),
            "clicks": num(get_cell(row, idx, "Clicks")),
            "cpc": num(get_cell(row, idx, "CPC")),
            "gmv_thb": num(get_cell(row, idx, "GMV")),
        }

    store_main_rows = build_sheet(store_ws, build_store_main)
    product_main_rows = build_sheet(store_ws, build_product_main)
    store_inner_rows = build_sheet(inner_ws, build_store_inner)
    product_inner_rows = build_sheet(inner_ws, build_product_inner)
    store_subsidy_rows = build_sheet(
        subsidy_ws,
        lambda row, idx: {
            "date": norm(get_cell(row, idx, "时间")),
            "store": norm(get_cell(row, idx, "Shop Name")),
            "actual_gmv": num(get_cell(row, idx, "GMV(After Seller Discounts)")),
            "platform_discount": num(get_cell(row, idx, "Platform Discount")),
            "voucher_from_shopee": num(get_cell(row, idx, "Voucher from shopee")),
        } if recognized(get_cell(row, idx, "时间")) and recognized(get_cell(row, idx, "Shop Name")) else None
    )

    rows = all_ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers) if header is not None}

    store_outer_rows = []
    store_brand_rows = []
    product_outer_rows = []
    product_brand_rows = []

    for row in rows:
        date_value = norm(get_cell(row, idx, "日期"))
        category_value = norm(get_cell(row, idx, "品类"))
        product = norm(get_cell(row, idx, "产品"))
        if not recognized(product):
            product = norm(get_cell(row, idx, "Product"))
        if not recognized(product):
            continue

        store = norm(get_cell(row, idx, "店铺"))
        category = category_value or category_store_map.get((store, product), "") or category_by_product.get(product, "")
        spend_usd = num(get_cell(row, idx, "花费金额（USD）"))
        impressions = num(get_cell(row, idx, "展示次数"))
        clicks = num(get_cell(row, idx, "点击量"))
        conversion_value = num(get_cell(row, idx, "购物转化价值"))
        orders = num(get_cell(row, idx, "订单数"))
        pitcher = norm(get_cell(row, idx, "投手"))
        ad_form = norm(get_cell(row, idx, "广告形式"))
        audience = norm(get_cell(row, idx, "广告人群"))
        ad_type = norm(get_cell(row, idx, "广告类型"))
        material_code = norm(get_cell(row, idx, "素材编号"))
        acquisition_type = norm(get_cell(row, idx, "拉新/再营销"))
        landing_page = norm(get_cell(row, idx, "投放页面"))
        ad_form2 = norm(get_cell(row, idx, "广告形式2"))

        store_item = {
            "date": date_value,
            "store": store,
            "category": category,
            "spend_usd": spend_usd,
            "impressions": impressions,
            "clicks": clicks,
            "conversion_value": conversion_value,
            "orders": orders,
            "ad_form": ad_form,
            "audience": audience,
            "ad_type": ad_type,
            "material_code": material_code,
            "acquisition_type": acquisition_type,
            "landing_page": landing_page,
            "ad_form2": ad_form2,
        }
        product_item = {
            "date": date_value,
            "store": store,
            "category": category,
            "product": product_store_map.get((store, product), product),
            "spend_usd": spend_usd,
            "impressions": impressions,
            "clicks": clicks,
            "conversion_value": conversion_value,
            "orders": orders,
            "ad_form": ad_form,
            "audience": audience,
            "ad_type": ad_type,
            "material_code": material_code,
            "acquisition_type": acquisition_type,
            "landing_page": landing_page,
            "ad_form2": ad_form2,
        }

        if pitcher == "SKT":
            store_brand_rows.append(store_item)
            product_brand_rows.append(product_item)
        else:
            store_outer_rows.append(store_item)
            product_outer_rows.append(product_item)

    return {
        "stores": sorted({row["store"] for row in store_main_rows if row.get("store")}),
        "main": store_main_rows,
        "inner": store_inner_rows,
        "subsidy": store_subsidy_rows,
        "outer": store_outer_rows,
        "brand": store_brand_rows,
    }, {
        "main": product_main_rows,
        "inner": product_inner_rows,
        "outer": product_outer_rows,
        "brand": product_brand_rows,
    }


def build_sp_tt_payload(wb):
    tt_ws = wb["TT出货"]
    sp_ws = wb["SP出货"]
    sales = defaultdict(lambda: {"ttSales": 0.0, "spSales": 0.0})

    def read_sheet(ws, sales_key, qty_header):
        rows = ws.iter_rows(min_row=1, values_only=True)
        headers = next(rows)
        idx = {header: i for i, header in enumerate(headers) if header is not None}
        for row in rows:
            dt = norm(get_cell(row, idx, "日期"))
            category = norm(get_cell(row, idx, "品类"))
            product_name = norm(get_cell(row, idx, "产品名称"))
            if not recognized(product_name):
                product_name = norm(get_cell(row, idx, "中文产品名"))
            if not recognized(dt) or not recognized(category) or not recognized(product_name):
                continue
            key = (dt, category, product_name)
            sales[key][sales_key] += num(get_cell(row, idx, qty_header))

    read_sheet(tt_ws, "ttSales", "TT销量")
    read_sheet(sp_ws, "spSales", "虾皮销量")

    rows = []
    for (dt, category, product_name), values in sorted(sales.items()):
        rows.append(
            {
                "date": dt,
                "category": category,
                "productName": product_name,
                "ttSales": values["ttSales"],
                "spSales": values["spSales"],
            }
        )
    return {"rows": rows}


def summarize_dates(rows):
    dates = sorted({row.get("date") for row in rows if row.get("date")})
    return {"count": len(rows), "min": dates[0] if dates else None, "max": dates[-1] if dates else None}


def write_manifest(path, payload):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    source_path = download_latest_workbook()
    wb = load_workbook(source_path)
    store_payload, product_payload = build_store_and_product_payloads(wb)
    sp_tt_payload = build_sp_tt_payload(wb)

    write_assignment(REPO_ROOT / "store-trend-data-stores.js", "window.storeTrendData = window.storeTrendData || {};", "window.storeTrendData.stores", store_payload["stores"])
    write_assignment(REPO_ROOT / "store-trend-data-main.js", "window.storeTrendData = window.storeTrendData || {};", "window.storeTrendData.main", store_payload["main"])
    write_assignment(REPO_ROOT / "store-trend-data-inner.js", "window.storeTrendData = window.storeTrendData || {};", "window.storeTrendData.inner", store_payload["inner"])
    write_assignment(REPO_ROOT / "store-trend-data-subsidy.js", "window.storeTrendData = window.storeTrendData || {};", "window.storeTrendData.subsidy", store_payload["subsidy"])
    write_chunked_loader(REPO_ROOT / "store-trend-data-outer.js", "window.storeTrendData = window.storeTrendData || {};", "window.storeTrendData.outer", store_payload["outer"])
    write_assignment(REPO_ROOT / "store-trend-data-brand.js", "window.storeTrendData = window.storeTrendData || {};", "window.storeTrendData.brand", store_payload["brand"])

    write_assignment(REPO_ROOT / "product-trend-data-main.js", "window.productTrendData = window.productTrendData || {};", "window.productTrendData.main", product_payload["main"])
    write_assignment(REPO_ROOT / "product-trend-data-inner.js", "window.productTrendData = window.productTrendData || {};", "window.productTrendData.inner", product_payload["inner"])
    write_chunked_loader(REPO_ROOT / "product-trend-data-outer.js", "window.productTrendData = window.productTrendData || {};", "window.productTrendData.outer", product_payload["outer"])
    write_assignment(REPO_ROOT / "product-trend-data-brand.js", "window.productTrendData = window.productTrendData || {};", "window.productTrendData.brand", product_payload["brand"])

    write_assignment(REPO_ROOT / "sp-tt-sales-data.js", "", "window.spTtSalesData", sp_tt_payload)

    write_manifest(
        REPO_ROOT / "dashboard-data-manifest.json",
        {
            "store_main": summarize_dates(store_payload["main"]),
            "store_inner": summarize_dates(store_payload["inner"]),
            "store_subsidy": summarize_dates(store_payload["subsidy"]),
            "store_outer": summarize_dates(store_payload["outer"]),
            "store_outer_parts": ["store-trend-data-outer.part1.js", "store-trend-data-outer.part2.js"],
            "store_brand": summarize_dates(store_payload["brand"]),
            "product_main": summarize_dates(product_payload["main"]),
            "product_inner": summarize_dates(product_payload["inner"]),
            "product_outer": summarize_dates(product_payload["outer"]),
            "product_outer_parts": ["product-trend-data-outer.part1.js", "product-trend-data-outer.part2.js"],
            "product_brand": summarize_dates(product_payload["brand"]),
            "sp_tt": summarize_dates(sp_tt_payload["rows"]),
        },
    )

    print(
        json.dumps(
            {
                "store_main": summarize_dates(store_payload["main"]),
                "store_inner": summarize_dates(store_payload["inner"]),
                "store_subsidy": summarize_dates(store_payload["subsidy"]),
                "store_outer": summarize_dates(store_payload["outer"]),
                "store_brand": summarize_dates(store_payload["brand"]),
                "product_main": summarize_dates(product_payload["main"]),
                "product_inner": summarize_dates(product_payload["inner"]),
                "product_outer": summarize_dates(product_payload["outer"]),
                "product_brand": summarize_dates(product_payload["brand"]),
                "sp_tt": summarize_dates(sp_tt_payload["rows"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
